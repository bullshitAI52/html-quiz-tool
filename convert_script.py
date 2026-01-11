import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import re

# Load source workbook
source_wb = openpyxl.load_workbook('低压.xlsx')

# Create target workbook
target_wb = Workbook()
ws = target_wb.active
ws.title = '题库'
ws.append(['类型', '题干', '选项A', '选项B', '选项C', '选项D', '答案', '解析'])

rows_to_process = 100 # Change this to limit how many rows we scan

def clean_text(text):
    if not text: return ""
    return str(text).strip()

# --- Process Choice Questions (选择题) ---
if '选择题' in source_wb.sheetnames:
    src_ws = source_wb['选择题']
    print("Processing Choice Questions...")
    
    current_q = None
    
    # Iterate through rows
    for i in range(1, rows_to_process + 1):
        cell = src_ws.cell(row=i, column=1)
        val = clean_text(cell.value)
        if not val: continue

        # Check if it's a question (starts with digit)
        if re.match(r'^\d+[、\.]', val):
            # Save previous question if exists
            if current_q:
                ws.append([
                    '单选', 
                    current_q['title'], 
                    current_q['A'], 
                    current_q['B'], 
                    current_q['C'], 
                    current_q['D'], 
                    current_q['content'], 
                    ''
                ])
            
            # Start new question
            current_q = {
                'title': val,
                'A': '', 'B': '', 'C': '', 'D': '',
                'content': '' # This will store the answer(s) like "A" or "AB"
            }
        
        # Check if it's an option (starts with A-D)
        elif re.match(r'^[A-D][.、]', val) and current_q:
            opt_letter = val[0] # 'A'
            opt_text = val[2:].strip() # Remove "A." portion
            current_q[opt_letter] = opt_text
            
            # Check for RED color (FFFF0000) indicating correct answer
            font = cell.font
            if font and font.color and font.color.rgb == 'FFFF0000':
                current_q['content'] += opt_letter

    # Append last question
    if current_q:
        ws.append([
            '单选', current_q['title'], current_q['A'], current_q['B'], current_q['C'], current_q['D'], current_q['content'], ''
        ])

# --- Process Judgment Questions (判断题) ---
# Logic: Read from specific sheets "判断题对" (True) and "判断题错" (False)
for sheet_name, answer in [('判断题对', '对'), ('判断题错', '错')]:
    if sheet_name in source_wb.sheetnames:
        print(f"Processing {sheet_name}...")
        src_ws = source_wb[sheet_name]
        for i in range(1, 20): # Scan first 20 rows of each judgment sheet
            cell = src_ws.cell(row=i, column=1)
            val = clean_text(cell.value)
            if not val or not re.match(r'^\d+[、\.]', val): continue
            
            ws.append(['判断', val, '', '', '', '', answer, ''])

# Save result
target_wb.save('低压_converted.xlsx')
print("Conversion complete: 低压_converted.xlsx")
